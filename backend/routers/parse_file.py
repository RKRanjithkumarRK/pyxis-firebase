"""
File parsing router — extract text from PDF, DOCX, XLSX, and plain text files.
Used for document Q&A: upload a file, ask questions about it.
"""

import io
import logging

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile

from core.auth import verify_token

logger = logging.getLogger(__name__)
router = APIRouter()

MAX_CHARS = 100_000


def _truncate(text: str) -> str:
    if len(text) <= MAX_CHARS:
        return text
    return text[:MAX_CHARS] + f"\n\n[Truncated — document exceeds {MAX_CHARS} characters]"


async def _parse_pdf(data: bytes) -> str:
    import pdfplumber
    text_parts = []
    with pdfplumber.open(io.BytesIO(data)) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if text:
                text_parts.append(text)
    return "\n\n".join(text_parts)


async def _parse_docx(data: bytes) -> str:
    from docx import Document
    doc = Document(io.BytesIO(data))
    return "\n".join(p.text for p in doc.paragraphs if p.text.strip())


async def _parse_xlsx(data: bytes) -> str:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(cells):
                rows.append(",".join(cells))
        if rows:
            parts.append(f"Sheet: {sheet_name}\n" + "\n".join(rows))
    wb.close()
    return "\n\n".join(parts)


@router.post("/parse-file")
async def parse_file(
    file: UploadFile = File(...),
    user: dict = Depends(verify_token),
):
    filename = file.filename or ""
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    data = await file.read()

    try:
        if ext == "pdf":
            text = await _parse_pdf(data)
        elif ext in ("docx", "doc"):
            text = await _parse_docx(data)
        elif ext in ("xlsx", "xls"):
            text = await _parse_xlsx(data)
        else:
            # Plain text fallback
            text = data.decode("utf-8", errors="replace")

        if not text.strip():
            raise HTTPException(status_code=422, detail="Could not extract text from file")

        return {"text": _truncate(text), "filename": filename, "chars": len(text)}

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"File parse error ({filename}): {e}")
        raise HTTPException(status_code=500, detail=f"Failed to parse file: {str(e)}")
